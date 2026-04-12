import os
import random
import json
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, make_response
)
import pymysql
import pymysql.cursors
import bcrypt
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.secret_key = os.getenv('SECRET_KEY', 'mabini-inventory-secret-key-change-me')

# ─── Database ────────────────────────────────────────────────────────────────

DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASS', ''),
    'database': os.getenv('DB_NAME', 'mabini_inventory'),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
}


def get_db():
    return pymysql.connect(**DB_CONFIG)


# ─── Auth helpers ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


# ─── Jinja helpers ───────────────────────────────────────────────────────────

@app.template_filter('number_format')
def number_format_filter(value, decimals=0):
    try:
        return f"{float(value):,.{decimals}f}"
    except (ValueError, TypeError):
        return value


@app.template_filter('date_format')
def date_format_filter(value, fmt='%b %d, %Y'):
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return value
    if isinstance(value, (datetime, date)):
        return value.strftime(fmt)
    return value


@app.context_processor
def inject_now():
    return {'now': datetime.now(), 'today': date.today()}


# ─── Routes ──────────────────────────────────────────────────────────────────

# ---------- Login / Logout ----------

@app.route('/')
@app.route('/index.php')
def index():
    return render_template('index.html')


@app.route('/login_process.php', methods=['POST'])
@app.route('/login', methods=['POST'])
def login_process():
    username = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()

    if not username or not password:
        flash('Please enter both username and password.', 'error')
        session['old_email'] = username
        return redirect(url_for('index'))

    try:
        conn = get_db()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, username, password, full_name, email, role, status "
                "FROM users WHERE (username=%s OR email=%s) AND status='Active'",
                (username, username),
            )
            user = cur.fetchone()
        conn.close()

        if user and bcrypt.checkpw(password.encode(), user['password'].encode()):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['user'] = user['full_name'] or user['username']
            session['email'] = user['email'] or ''
            session['role'] = user['role']
            session['logged_in'] = True

            if request.form.get('remember'):
                resp = make_response(redirect(url_for('dashboard')))
                resp.set_cookie('remember_email', username, max_age=30*86400)
                return resp

            return redirect(url_for('dashboard'))

        flash('Invalid username or password.', 'error')
        session['old_email'] = username
        return redirect(url_for('index'))

    except Exception:
        flash('Database error. Please try again later.', 'error')
        session['old_email'] = username
        return redirect(url_for('index'))


@app.route('/logout.php', methods=['GET', 'POST'])
@app.route('/logout', methods=['GET', 'POST'])
def logout():
    session.clear()
    flash('You have been successfully logged out.', 'success')
    return redirect(url_for('index'))


# ---------- Dashboard ----------

@app.route('/dashboard.php')
@app.route('/dashboard')
@login_required
def dashboard():
    # Session-based stats (legacy pages)
    products = session.get('products', [])
    costs = session.get('costs', [])
    quantities = session.get('quantities', [])
    issued = session.get('issued', [])

    total_products = len(products)
    total_quantity = sum(float(q.get('quantity', 0)) for q in quantities)
    total_issued = sum(float(i.get('quantity_issued', i.get('quantity', 0))) for i in issued)
    total_value = sum(float(c.get('unit_cost', 0)) for c in costs)

    return render_template('dashboard.html',
                           total_products=total_products,
                           total_quantity=total_quantity,
                           total_issued=total_issued,
                           total_value=total_value)


# ---------- Item Master List ----------

@app.route('/item-master-list.php', methods=['GET', 'POST'])
@app.route('/item-master-list', methods=['GET', 'POST'])
@login_required
def item_master_list():
    conn = get_db()

    # AJAX item details
    if request.args.get('action') == 'get_item_details' and request.args.get('id'):
        item_id = int(request.args['id'])
        with conn.cursor() as cur:
            cur.execute(
                "SELECT i.*, c.category_name FROM items i "
                "LEFT JOIN categories c ON i.category_id=c.id WHERE i.id=%s",
                (item_id,),
            )
            item = cur.fetchone()
            if not item:
                conn.close()
                return jsonify(error='Item not found')
            # Convert date/datetime values to strings for JSON
            for k, v in item.items():
                if isinstance(v, (datetime, date)):
                    item[k] = v.isoformat()

            cur.execute(
                "SELECT DISTINCT o.office_name, it.transaction_date, "
                "COALESCE(NULLIF(it.quantity,0), it.quantity_approved, it.quantity_requested, 0) AS quantity, "
                "it.status "
                "FROM inventory_transactions it JOIN offices o ON it.office_id=o.id "
                "WHERE it.item_id=%s AND it.transaction_type='ISSUE' "
                "ORDER BY it.transaction_date DESC LIMIT 5",
                (item_id,),
            )
            agencies = cur.fetchall()
            for a in agencies:
                for k, v in a.items():
                    if isinstance(v, (datetime, date)):
                        a[k] = v.isoformat()
        conn.close()
        return jsonify(item=item, agencies=agencies)

    # POST actions
    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'add_item':
            exp = request.form.get('expiration_date') or None
            acq = request.form.get('date_acquired') or None
            office_id = request.form.get('office_id') or None
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO items (item_code,item_name,category_id,unit,unit_cost,"
                    "quantity_on_hand,reorder_level,expiration_date,date_acquired,office_id,status,created_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Active',%s)",
                    (
                        request.form.get('item_code', ''),
                        request.form.get('item_name', ''),
                        request.form.get('category') or None,
                        request.form.get('unit', 'piece'),
                        request.form.get('unit_cost', 0),
                        request.form.get('quantity_on_hand', 0),
                        request.form.get('reorder_level', 10),
                        exp,
                        acq,
                        office_id,
                        session.get('user_id'),
                    ),
                )
            conn.commit()
            flash('Item added successfully!', 'success')

        elif action == 'update_item':
            exp = request.form.get('expiration_date') or None
            acq = request.form.get('date_acquired') or None
            office_id = request.form.get('office_id') or None
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE items SET item_code=%s,item_name=%s,category_id=%s,unit=%s,"
                    "unit_cost=%s,quantity_on_hand=%s,reorder_level=%s,expiration_date=%s,"
                    "date_acquired=%s,office_id=%s,status=%s,updated_at=NOW() WHERE id=%s",
                    (
                        request.form.get('item_code', ''),
                        request.form.get('item_name', ''),
                        request.form.get('category') or None,
                        request.form.get('unit', 'piece'),
                        request.form.get('unit_cost', 0),
                        request.form.get('quantity_on_hand', 0),
                        request.form.get('reorder_level', 10),
                        exp,
                        acq,
                        office_id,
                        request.form.get('status', 'Active'),
                        request.form.get('id', 0),
                    ),
                )
            conn.commit()
            flash('Item updated successfully!', 'success')

        elif action == 'delete_item':
            with conn.cursor() as cur:
                cur.execute("DELETE FROM items WHERE id=%s", (request.form.get('id', 0),))
            conn.commit()
            flash('Item deleted successfully!', 'success')

        elif action == 'upload_excel':
            f = request.files.get('excel_file')
            if f:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(f.read()))
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                imported = 0
                skipped = 0
                category_cache = {}  # name -> id
                office_cache = {}  # name -> id

                duplicates = 0
                with conn.cursor() as cur:
                    # Pre-load existing categories
                    cur.execute("SELECT id, category_name FROM categories")
                    for cat in cur.fetchall():
                        category_cache[cat['category_name'].strip().lower()] = cat['id']

                    # Get max auto-generated code number
                    cur.execute("SELECT item_code FROM items WHERE item_code LIKE 'ITEM-%%' ORDER BY item_code DESC LIMIT 1")
                    last_auto = cur.fetchone()
                    auto_num = 1
                    if last_auto:
                        try:
                            auto_num = int(last_auto['item_code'].split('-')[1]) + 1
                        except (IndexError, ValueError):
                            pass

                    for row in rows:
                        if len(row) < 2 or (not row[0] and not row[1]):
                            skipped += 1
                            continue

                        item_code = str(row[0]).strip() if row[0] else ''
                        item_name = str(row[1]).strip() if row[1] else ''
                        cat_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                        unit = str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'piece'
                        unit_cost = float(row[4]) if len(row) > 4 and row[4] else 0
                        quantity = int(float(str(row[5]))) if len(row) > 5 and row[5] else 0
                        reorder = int(float(str(row[6]))) if len(row) > 6 and row[6] else 10
                        exp_date = None
                        if len(row) > 7 and row[7]:
                            if isinstance(row[7], (datetime, date)):
                                exp_date = row[7].strftime('%Y-%m-%d')
                            else:
                                exp_date = str(row[7]).strip() or None

                        acq_date = None
                        if len(row) > 8 and row[8]:
                            if isinstance(row[8], (datetime, date)):
                                acq_date = row[8].strftime('%Y-%m-%d')
                            else:
                                acq_date = str(row[8]).strip() or None

                        office_name = str(row[9]).strip() if len(row) > 9 and row[9] else ''

                        # Auto-generate item code if empty
                        if not item_code:
                            item_code = f"ITEM-{auto_num:05d}"
                            auto_num += 1

                        # Auto-match or create category
                        category_id = None
                        if cat_name:
                            key = cat_name.lower()
                            if key in category_cache:
                                category_id = category_cache[key]
                            else:
                                cur.execute(
                                    "INSERT INTO categories (category_name, description, status) VALUES (%s, %s, 'Active')",
                                    (cat_name, f'Auto-created from Excel import'),
                                )
                                conn.commit()
                                category_id = cur.lastrowid
                                category_cache[key] = category_id

                        # Auto-match office by name
                        matched_office_id = None
                        if office_name:
                            okey = office_name.lower()
                            if okey not in office_cache:
                                cur.execute("SELECT id FROM offices WHERE LOWER(office_name)=%s AND status='Active'", (okey,))
                                orow = cur.fetchone()
                                if orow:
                                    office_cache[okey] = orow['id']
                            matched_office_id = office_cache.get(okey)

                        try:
                            cur.execute(
                                "INSERT INTO items (item_code,item_name,category_id,unit,unit_cost,"
                                "quantity_on_hand,reorder_level,expiration_date,date_acquired,office_id,status,created_by) "
                                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Active',%s)",
                                (
                                    item_code, item_name, category_id,
                                    unit, unit_cost, quantity, reorder, exp_date, acq_date,
                                    matched_office_id,
                                    session.get('user_id'),
                                ),
                            )
                            imported += 1
                        except pymysql.err.IntegrityError:
                            duplicates += 1
                conn.commit()
                msg = f'Successfully imported {imported} items!'
                if skipped:
                    msg += f' ({skipped} empty rows skipped)'
                if duplicates:
                    msg += f' ({duplicates} duplicates skipped)'
                flash(msg, 'success')

        conn.close()
        return redirect(url_for('item_master_list'))

    # GET
    with conn.cursor() as cur:
        cur.execute(
            "SELECT i.*, c.category_name, o.office_name FROM items i "
            "LEFT JOIN categories c ON i.category_id=c.id "
            "LEFT JOIN offices o ON i.office_id=o.id "
            "ORDER BY i.item_code ASC"
        )
        items = cur.fetchall()
        cur.execute("SELECT id, category_name FROM categories WHERE status='Active' ORDER BY category_name")
        categories = cur.fetchall()
        cur.execute("SELECT id, office_name FROM offices WHERE status='Active' ORDER BY office_name")
        offices_list = cur.fetchall()
    conn.close()

    total_items = len(items)
    low_stock = 0
    total_value = 0.0
    for it in items:
        qty = float(it.get('quantity_on_hand') or 0)
        reorder = float(it.get('reorder_level') or 0)
        cost = float(it.get('unit_cost') or 0)
        if qty <= reorder and reorder > 0:
            low_stock += 1
        total_value += qty * cost

    return render_template('item_master_list.html',
                           items=items, categories=categories, offices=offices_list,
                           total_items=total_items, low_stock=low_stock,
                           total_value=total_value)


# ---------- Export Items to Excel ----------

@app.route('/export-items')
@login_required
def export_items():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    conn = get_db()
    with conn.cursor() as cur:
        cur.execute(
            "SELECT i.item_code, i.item_name, c.category_name, o.office_name, i.unit, i.unit_cost, "
            "i.quantity_on_hand, i.reorder_level, i.expiration_date, i.date_acquired "
            "FROM items i LEFT JOIN categories c ON i.category_id=c.id "
            "LEFT JOIN offices o ON i.office_id=o.id ORDER BY i.item_code ASC"
        )
        items = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Items'

    headers = ['Item Code', 'Item Name', 'Category', 'Office', 'Unit', 'Unit Cost', 'Quantity', 'Reorder Level', 'Expiration Date', 'Date Acquired']
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    widths = [15, 35, 25, 25, 12, 12, 12, 15, 18, 18]
    for i, w in enumerate(widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
        ws.column_dimensions[col_letter].width = w

    data_font = Font(name='Arial', size=10)
    for r, item in enumerate(items, 2):
        values = [
            item['item_code'], item['item_name'], item.get('category_name', ''),
            item.get('office_name', ''),
            item['unit'], float(item.get('unit_cost') or 0),
            int(item.get('quantity_on_hand') or 0), int(item.get('reorder_level') or 0),
            item['expiration_date'].strftime('%Y-%m-%d') if item.get('expiration_date') else '',
            item['date_acquired'].strftime('%Y-%m-%d') if item.get('date_acquired') else '',
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val if val != '' else None)
            cell.font = data_font
            cell.border = thin_border
            if c == 6:
                cell.number_format = '#,##0.00'
            if c in (7, 8):
                cell.number_format = '#,##0'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    resp = make_response(output.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=inventory_items_{timestamp}.xlsx'
    return resp


# ---------- Issue Items ----------

@app.route('/issue-items.php', methods=['GET', 'POST'])
@app.route('/issue-items', methods=['GET', 'POST'])
@login_required
def issue_items():
    conn = get_db()

    if request.method == 'POST' and request.form.get('action') == 'create_issue':
        try:
            item_id = request.form['item_id']
            office_id = request.form['office_id']
            quantity_requested = int(request.form['quantity_requested'])
            remarks = request.form.get('remarks', '')

            with conn.cursor() as cur:
<<<<<<< HEAD
                if action == 'create_issue':
                    item_id = request.form['item_id']
                    office_id = request.form['office_id']
                    quantity_requested = int(request.form['quantity_requested'])
                    remarks = request.form.get('remarks', '')

                    cur.execute("SELECT item_code,item_name,quantity_on_hand,unit_cost FROM items WHERE id=%s", (item_id,))
                    item = cur.fetchone()
                    if not item:
                        raise Exception('Item not found')

                    # Don't block if insufficient - supervisor will approve what's available.
                    stock_warning = ''
                    if item['quantity_on_hand'] < quantity_requested:
                        stock_warning = f" (Note: Only {item['quantity_on_hand']} available in stock)"

                    ref = f"ISS-{datetime.now():%Y%m%d}-{random.randint(1,9999):04d}"
                    total_cost = quantity_requested * float(item['unit_cost'])
                    cur.execute(
                        "INSERT INTO inventory_transactions "
                        "(transaction_type,transaction_date,reference_number,office_id,item_id,"
                        "quantity,quantity_requested,unit_cost,total_cost,remarks,created_by,status) "
                        "VALUES ('ISSUE',CURDATE(),%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pending')",
                        (
                            ref,
                            office_id,
                            item_id,
                            quantity_requested,
                            quantity_requested,
                            item['unit_cost'],
                            total_cost,
                            remarks,
                            user_id,
                        ),
                    )
                    flash(f"Issue request created successfully! Reference: {ref}{stock_warning}", 'success')

                elif action == 'approve':
                    tid = request.form.get('transaction_id')
                    qty_approved_raw = request.form.get('quantity_approved')
                    if not tid:
                        raise Exception('Missing transaction id')

                    if qty_approved_raw:
                        qty_approved = int(qty_approved_raw)
                        cur.execute(
                            "UPDATE inventory_transactions SET status='Approved', "
                            "quantity_approved=%s,quantity=%s,processed_by=%s,updated_at=NOW() "
                            "WHERE id=%s AND transaction_type='ISSUE'",
                            (qty_approved, qty_approved, user_id, tid),
                        )
                        flash(f'Transaction approved with {qty_approved} units!', 'success')
                    else:
                        cur.execute(
                            "UPDATE inventory_transactions SET status='Approved', "
                            "processed_by=%s,updated_at=NOW() "
                            "WHERE id=%s AND transaction_type='ISSUE'",
                            (user_id, tid),
                        )
                        flash('Transaction approved successfully!', 'success')

                elif action == 'complete':
                    tid = request.form.get('transaction_id')
                    if not tid:
                        raise Exception('Missing transaction id')

                    cur.execute(
                        "SELECT item_id, quantity, quantity_approved, transaction_type, status "
                        "FROM inventory_transactions WHERE id=%s",
                        (tid,),
                    )
                    trans = cur.fetchone()
                    if not trans or trans['transaction_type'] != 'ISSUE':
                        raise Exception('Issue transaction not found')
                    if trans['status'] != 'Approved':
                        raise Exception('Only approved transactions can be completed')

                    qty_to_issue = trans.get('quantity_approved') or trans['quantity']
                    cur.execute("SELECT quantity_on_hand FROM items WHERE id=%s", (trans['item_id'],))
                    item = cur.fetchone()
                    if not item:
                        raise Exception('Source item not found in warehouse')
                    if item['quantity_on_hand'] < qty_to_issue:
                        raise Exception('Insufficient stock to complete this transaction')

                    cur.execute(
                        "UPDATE items SET quantity_on_hand=quantity_on_hand-%s, updated_at=NOW() "
                        "WHERE id=%s",
                        (qty_to_issue, trans['item_id']),
                    )
                    cur.execute("SELECT quantity_on_hand FROM items WHERE id=%s", (trans['item_id'],))
                    new_bal = cur.fetchone()['quantity_on_hand']
                    cur.execute(
                        "INSERT INTO stock_movements "
                        "(item_id, transaction_id, movement_type, quantity, balance_after, created_by) "
                        "VALUES (%s, %s, 'OUT', %s, %s, %s)",
                        (trans['item_id'], tid, qty_to_issue, new_bal, user_id),
                    )
                    cur.execute(
                        "UPDATE inventory_transactions SET status='Completed', processed_by=%s, updated_at=NOW() "
                        "WHERE id=%s",
                        (user_id, tid),
                    )
                    flash('Transaction completed successfully! Items transferred to office.', 'success')

                elif action == 'cancel':
                    tid = request.form.get('transaction_id')
                    if not tid:
                        raise Exception('Missing transaction id')
                    cur.execute(
                        "UPDATE inventory_transactions SET status='Cancelled', processed_by=%s, updated_at=NOW() "
                        "WHERE id=%s AND transaction_type='ISSUE'",
                        (user_id, tid),
                    )
                    flash('Transaction cancelled', 'success')

                else:
                    raise Exception('Invalid action')

=======
                cur.execute(
                    "SELECT item_code,item_name,quantity_on_hand,unit_cost,office_id "
                    "FROM items WHERE id=%s",
                    (item_id,)
                )
                item = cur.fetchone()
                if not item:
                    raise Exception('Item not found')
                if item.get('office_id') not in (None, 0):
                    raise Exception('Selected item is office-assigned. Please choose a warehouse item.')
                # Don't block if insufficient - supervisor will approve what's available
                # Just warn in flash message
                stock_warning = ''
                if item['quantity_on_hand'] < quantity_requested:
                    stock_warning = f" (Note: Only {item['quantity_on_hand']} available in stock)"

                ref = f"ISS-{datetime.now():%Y%m%d}-{random.randint(1,9999):04d}"
                txn_no = f"TXN-{datetime.now():%Y%m%d%H%M%S%f}-{random.randint(100,999)}"
                total_cost = quantity_requested * float(item['unit_cost'])
                user_id = session.get('user_id', 1)
                cur.execute(
                    "INSERT INTO inventory_transactions "
                    "(transaction_number,transaction_type,transaction_date,reference_number,office_id,item_id,"
                    "quantity,quantity_requested,unit_cost,total_cost,remarks,created_by,status) "
                    "VALUES (%s,'ISSUE',CURDATE(),%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pending')",
                    (txn_no, ref, office_id, item_id, quantity_requested, quantity_requested,
                     item['unit_cost'], total_cost, remarks, user_id),
                )
>>>>>>> 5938f7e7c272e1411a5ffe1a94a86d54b7cb137b
            conn.commit()
            flash(f"Issue request created successfully! Reference: {ref}{stock_warning}", 'success')
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}", 'error')
        conn.close()
        return redirect(url_for('issue_items'))

    with conn.cursor() as cur:
        cur.execute(
            "SELECT it.*,o.office_name,i.item_name,i.item_code,u.full_name as processed_by_name "
            "FROM inventory_transactions it "
            "LEFT JOIN offices o ON it.office_id=o.id "
            "LEFT JOIN items i ON it.item_id=i.id "
            "LEFT JOIN users u ON it.processed_by=u.id "
            "WHERE it.transaction_type='ISSUE' ORDER BY it.created_at DESC LIMIT 50"
        )
        transactions = cur.fetchall()
        cur.execute("SELECT id,office_name FROM offices WHERE status='Active' ORDER BY office_name")
        offices = cur.fetchall()
        cur.execute(
            "SELECT id,item_code,item_name,quantity_on_hand,unit "
            "FROM items "
            "WHERE status='Active' AND (office_id IS NULL OR office_id=0) "
            "ORDER BY item_name"
        )
        items = cur.fetchall()
    conn.close()

    today_str = date.today().isoformat()
    pending_count = sum(1 for t in transactions if t['status'] == 'Pending')
    completed_today = sum(1 for t in transactions if t['status'] == 'Completed' and str(t.get('transaction_date', '')) == today_str)

    return render_template('issue_items.html', transactions=transactions, offices=offices, items=items,
                           pending_count=pending_count, completed_today=completed_today)


# ---------- Receive Items ----------

@app.route('/receive-items.php', methods=['GET', 'POST'])
@app.route('/receive-items', methods=['GET', 'POST'])
@login_required
def receive_items():
    conn = get_db()

    if request.method == 'POST' and request.form.get('action') == 'receive_items':
        try:
            item_id = request.form['item_id']
            quantity = int(request.form['quantity'])
            unit_cost = float(request.form['unit_cost'])
            supplier_id = request.form.get('supplier_id') or None
            remarks = request.form.get('remarks', '')

            with conn.cursor() as cur:
                cur.execute("SELECT item_code,item_name,quantity_on_hand FROM items WHERE id=%s", (item_id,))
                item = cur.fetchone()
                if not item:
                    raise Exception('Item not found')

                ref = f"RCV-{datetime.now():%Y%m%d}-{random.randint(1,9999):04d}"
                txn_no = f"TXN-{datetime.now():%Y%m%d%H%M%S%f}-{random.randint(100,999)}"
                total_cost = quantity * unit_cost
                user_id = session.get('user_id', 1)
                cur.execute(
                    "INSERT INTO inventory_transactions "
                    "(transaction_number,transaction_type,transaction_date,reference_number,item_id,"
                    "quantity,unit_cost,total_cost,remarks,created_by,processed_by,status) "
                    "VALUES (%s,'RECEIVE',CURDATE(),%s,%s,%s,%s,%s,%s,%s,%s,'Completed')",
                    (txn_no, ref, item_id, quantity, unit_cost, total_cost, remarks, user_id, user_id),
                )
                new_qty = item['quantity_on_hand'] + quantity
                cur.execute(
                    "UPDATE items SET quantity_on_hand=%s,unit_cost=%s,"
                    "supplier_id=COALESCE(%s,supplier_id),updated_at=NOW() WHERE id=%s",
                    (new_qty, unit_cost, supplier_id, item_id),
                )
                cur.execute(
                    "INSERT INTO stock_movements "
                    "(item_id,transaction_id,movement_type,quantity,balance_after,reference,remarks,created_by) "
                    "VALUES (%s,LAST_INSERT_ID(),'IN',%s,%s,%s,%s,%s)",
                    (item_id, quantity, new_qty, ref, remarks, user_id),
                )
            conn.commit()
            flash(f"Items received successfully! Reference: {ref}", 'success')
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}", 'error')
        conn.close()
        return redirect(url_for('receive_items'))

    with conn.cursor() as cur:
        cur.execute(
            "SELECT it.*,i.item_name,i.item_code,u.full_name as processed_by_name "
            "FROM inventory_transactions it LEFT JOIN items i ON it.item_id=i.id "
            "LEFT JOIN users u ON it.processed_by=u.id "
            "WHERE it.transaction_type='RECEIVE' ORDER BY it.created_at DESC LIMIT 50"
        )
        receipts = cur.fetchall()
        cur.execute("SELECT id,item_code,item_name,unit_cost,unit FROM items WHERE status='Active' ORDER BY item_name")
        items = cur.fetchall()
        
        # Try to get suppliers, but handle if table doesn't exist
        try:
            cur.execute("SELECT id,supplier_name FROM suppliers WHERE status='Active' ORDER BY supplier_name")
            suppliers = cur.fetchall()
        except:
            suppliers = []
    conn.close()

    today_str = date.today().isoformat()
    received_today = sum(1 for r in receipts if str(r.get('transaction_date', '')) == today_str)
    received_month = sum(1 for r in receipts if str(r.get('transaction_date', ''))[:7] == datetime.now().strftime('%Y-%m'))
    total_value = sum(float(r.get('total_cost', 0) or 0) for r in receipts)

    return render_template('receive_items.html', receipts=receipts, items=items, suppliers=suppliers,
                           received_today=received_today, received_month=received_month, total_value=total_value)


# ---------- Process Transactions ----------

@app.route('/process-transactions.php', methods=['GET', 'POST'])
@app.route('/process-transactions', methods=['GET', 'POST'])
@login_required
def process_transactions():
    conn = get_db()

    if request.method == 'POST' and request.form.get('action'):
        action = request.form['action']
        tid = request.form['transaction_id']
        user_id = session.get('user_id', 1)

        try:
            with conn.cursor() as cur:
                if action == 'approve':
                    quantity_approved = request.form.get('quantity_approved')
                    if quantity_approved:
                        # Update with approved quantity
                        quantity_approved = int(quantity_approved)
                        # Also update the quantity field to match approved quantity
                        cur.execute(
                            "UPDATE inventory_transactions SET status='Approved',"
                            "quantity_approved=%s,quantity=%s,processed_by=%s,updated_at=NOW() WHERE id=%s",
                            (quantity_approved, quantity_approved, user_id, tid)
                        )
                        flash(f'Transaction approved with {quantity_approved} units!', 'success')
                    else:
                        # Simple approval without quantity modification
                        cur.execute(
                            "UPDATE inventory_transactions SET status='Approved',"
                            "processed_by=%s,updated_at=NOW() WHERE id=%s",
                            (user_id, tid)
                        )
                        flash('Transaction approved successfully!', 'success')

                elif action == 'complete':
                    # Get transaction details including office_id and quantity_approved
                    cur.execute(
                        "SELECT item_id, quantity, quantity_requested, quantity_approved, transaction_type, office_id "
                        "FROM inventory_transactions WHERE id=%s", 
                        (tid,)
                    )
                    trans = cur.fetchone()
                    if not trans:
                        raise Exception('Transaction not found')

                    completed_qty = int(trans.get('quantity') or 0)
                    
                    if trans['transaction_type'] == 'ISSUE':
                        # For legacy rows, quantity can be 0 while requested/approved is set.
                        qty_to_issue = trans.get('quantity_approved') or trans.get('quantity') or trans.get('quantity_requested') or 0
                        qty_to_issue = int(qty_to_issue)
                        if qty_to_issue <= 0:
                            raise Exception('Cannot complete issue transaction with zero quantity')
                        completed_qty = qty_to_issue
                        
                        # 1. Verify warehouse has the item
                        cur.execute(
                            "SELECT quantity_on_hand FROM items "
                            "WHERE id=%s AND (office_id IS NULL OR office_id=0)",
                            (trans['item_id'],)
                        )
                        warehouse_item = cur.fetchone()
                        if not warehouse_item:
                            raise Exception('Source item not found in warehouse')
                        
                        if warehouse_item['quantity_on_hand'] < qty_to_issue:
                            raise Exception('Insufficient stock to complete this transaction')
                        
                        # 2. Decrement warehouse stock
                        cur.execute(
                            "UPDATE items SET quantity_on_hand=quantity_on_hand-%s, updated_at=NOW() "
                            "WHERE id=%s",
                            (qty_to_issue, trans['item_id']),
                        )
                        
                        # 3. Get new balance and log stock movement OUT from warehouse
                        cur.execute("SELECT quantity_on_hand FROM items WHERE id=%s", (trans['item_id'],))
                        new_bal = cur.fetchone()['quantity_on_hand']
                        cur.execute(
                            "INSERT INTO stock_movements (item_id, transaction_id, movement_type, quantity, balance_after, created_by) "
                            "VALUES (%s, %s, 'OUT', %s, %s, %s)",
                            (trans['item_id'], tid, qty_to_issue, new_bal, user_id),
                        )
                    
                    # Mark transaction as completed
                    cur.execute(
                        "UPDATE inventory_transactions SET status='Completed', quantity=%s, processed_by=%s, updated_at=NOW() WHERE id=%s", 
                        (completed_qty, user_id, tid)
                    )
                    flash('Transaction completed successfully! Items transferred to office.', 'success')

                elif action == 'cancel':
                    cur.execute("UPDATE inventory_transactions SET status='Cancelled',processed_by=%s,updated_at=NOW() WHERE id=%s", (user_id, tid))
                    flash('Transaction cancelled', 'success')

            conn.commit()
        except Exception as e:
            conn.rollback()
            flash(f"Error: {e}", 'error')
        conn.close()
        return redirect(url_for('process_transactions'))

    status_filter = request.args.get('status', 'all')
    type_filter = request.args.get('type', 'all')
    if type_filter != 'all':
        type_filter = type_filter.upper()

    query = (
        "SELECT it.*,o.office_name,i.item_name,i.item_code,i.quantity_on_hand,"
        "u.full_name as processed_by_name FROM inventory_transactions it "
        "LEFT JOIN offices o ON it.office_id=o.id "
        "LEFT JOIN items i ON it.item_id=i.id "
        "LEFT JOIN users u ON it.processed_by=u.id WHERE 1=1 "
    )
    params = []
    if status_filter != 'all':
        query += " AND it.status=%s"
        params.append(status_filter)
    if type_filter != 'all':
        query += " AND it.transaction_type=%s"
        params.append(type_filter)
    query += (
        " ORDER BY CASE it.status WHEN 'Pending' THEN 1 WHEN 'Approved' THEN 2 "
        "WHEN 'Completed' THEN 3 ELSE 4 END, it.created_at DESC LIMIT 100"
    )

    with conn.cursor() as cur:
        cur.execute(query, params)
        transactions = cur.fetchall()
    conn.close()

    today_str = date.today().isoformat()
    pending_count = sum(1 for t in transactions if t['status'] == 'Pending')
    approved_count = sum(1 for t in transactions if t['status'] == 'Approved')
    completed_today = sum(1 for t in transactions if t['status'] == 'Completed' and str(t.get('updated_at', ''))[:10] == today_str)

    return render_template('process_transactions.html',
                           transactions=transactions,
                           status_filter=status_filter,
                           type_filter=type_filter,
                           pending_count=pending_count,
                           approved_count=approved_count,
                           completed_today=completed_today)


# ---------- Offices ----------

@app.route('/offices.php', methods=['GET', 'POST'])
@app.route('/offices', methods=['GET', 'POST'])
@login_required
def offices():
    conn = get_db()

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add_office':
            name = request.form.get('office_name', '').strip()
            if name:
                code = name.upper().replace(' ', '')[:20]
                with conn.cursor() as cur:
                    cur.execute(
                        "INSERT INTO offices (office_code, office_name, status) VALUES (%s, %s, 'Active')",
                        (code, name),
                    )
                conn.commit()
            flash('Office added successfully!', 'success')
        elif action == 'edit_office':
            oid = int(request.form.get('id', 0))
            new_name = request.form.get('office_name', '').strip()
            if new_name:
                with conn.cursor() as cur:
                    cur.execute("UPDATE offices SET office_name=%s, updated_at=NOW() WHERE id=%s", (new_name, oid))
                conn.commit()
            flash('Office updated successfully!', 'success')
        elif action == 'delete_office':
            oid = int(request.form.get('id', 0))
            with conn.cursor() as cur:
                cur.execute("DELETE FROM offices WHERE id=%s", (oid,))
            conn.commit()
            flash('Office deleted successfully!', 'success')
        conn.close()
        return redirect(url_for('offices'))

    with conn.cursor() as cur:
        cur.execute("SELECT * FROM offices ORDER BY office_name")
        offices_list = cur.fetchall()
    conn.close()

    return render_template('offices.html', offices=offices_list)


@app.route('/offices/items/<int:office_id>')
@login_required
def office_items(office_id):
    """
    Show items issued to an office by querying completed Issue transactions.
    Each row represents an issue transaction (not aggregated, so you can see history).
    """
    conn = get_db()
    with conn.cursor() as cur:
        # Query completed Issue transactions for this office
        cur.execute("""
            SELECT i.item_code, i.item_name, c.category_name, o.office_name,
                   i.unit, 
                   COALESCE(it.quantity_approved, it.quantity) as quantity_issued,
                   i.unit_cost, 
                   it.updated_at as date_issued,
                   it.reference_number
            FROM inventory_transactions it
            INNER JOIN items i ON it.item_id = i.id
            LEFT JOIN categories c ON i.category_id = c.id
            LEFT JOIN offices o ON it.office_id = o.id
            WHERE it.office_id = %s 
              AND it.transaction_type = 'ISSUE' 
              AND it.status = 'Completed'
            ORDER BY it.updated_at DESC, i.item_code
        """, (office_id,))
        transactions = cur.fetchall()
    conn.close()
    
    result = []
    for trans in transactions:
        qty = float(trans.get('quantity_issued', 0) or 0)
        cost = float(trans.get('unit_cost', 0) or 0)
        result.append({
            'item_code': trans.get('item_code', ''),
            'item_name': trans.get('item_name', ''),
            'category_name': trans.get('category_name', ''),
            'office_name': trans.get('office_name', ''),
            'unit': trans.get('unit', ''),
            'quantity_on_hand': qty,  # Keep same field name for frontend compatibility
            'unit_cost': cost,
            'total_cost': round(qty * cost, 2),
            'date_acquired': str(trans['date_issued']) if trans.get('date_issued') else '',
            'reference': trans.get('reference_number', '')
        })
    return jsonify(result)


@app.route('/offices/export/<int:office_id>')
@login_required
def export_office_items(office_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    # Get optional date filter from query parameter
    date_filter = request.args.get('date', None)

    conn = get_db()
    with conn.cursor() as cur:
        cur.execute("SELECT office_name FROM offices WHERE id=%s", (office_id,))
        office = cur.fetchone()
        
        # Build query with optional date filter
        if date_filter:
            cur.execute("""
                SELECT i.item_code, i.item_name, c.category_name, o.office_name,
                       i.unit, i.quantity_on_hand, i.unit_cost, i.date_acquired
                FROM items i
                LEFT JOIN categories c ON i.category_id = c.id
                LEFT JOIN offices o ON i.office_id = o.id
                WHERE i.office_id = %s AND i.date_acquired = %s
                ORDER BY i.item_code
            """, (office_id, date_filter))
        else:
            cur.execute("""
                SELECT i.item_code, i.item_name, c.category_name, o.office_name,
                       i.unit, i.quantity_on_hand, i.unit_cost, i.date_acquired
                FROM items i
                LEFT JOIN categories c ON i.category_id = c.id
                LEFT JOIN offices o ON i.office_id = o.id
                WHERE i.office_id = %s
                ORDER BY i.item_code
            """, (office_id,))
        items = cur.fetchall()
    conn.close()

    office_name = office['office_name'] if office else 'Unknown'

    wb = Workbook()
    ws = wb.active
    ws.title = office_name[:31]

    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['No.', 'Item Code', 'Item Name', 'Category', 'Office', 'Unit',
               'Quantity', 'Unit Cost', 'Total Cost', 'Date Acquired']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    widths = [6, 15, 35, 25, 25, 10, 12, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    data_font = Font(name='Arial', size=10)
    for idx, item in enumerate(items, 1):
        qty = float(item.get('quantity_on_hand', 0) or 0)
        cost = float(item.get('unit_cost', 0) or 0)
        row = [
            idx,
            item.get('item_code', ''),
            item.get('item_name', ''),
            item.get('category_name', ''),
            item.get('office_name', ''),
            item.get('unit', ''),
            qty,
            cost,
            round(qty * cost, 2),
            str(item['date_acquired']) if item.get('date_acquired') else ''
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c in (8, 9):
                cell.number_format = '#,##0.00'
            if c == 7:
                cell.number_format = '#,##0'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = office_name.replace(' ', '_')
    # Add date to filename if filtered
    if date_filter:
        filename = f'{safe_name}_Items_{date_filter}.xlsx'
    else:
        filename = f'{safe_name}_Items.xlsx'
    
    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


# ---------- Report ----------

@app.route('/report.php')
@app.route('/report')
@login_required
def report():
    products = session.get('products', [])
    offices_list = session.get('offices', [])
    issued = session.get('issued', [])

    total_items = len(products)
    total_offices = len(offices_list)
    total_issued = sum(float(i.get('quantity_issued', i.get('quantity', 0))) for i in issued)
    total_value = sum(float(p.get('price', 0)) for p in products)

    return render_template('report.html',
                           products=products, total_items=total_items,
                           total_offices=total_offices, total_issued=total_issued,
                           total_value=total_value)


# ---------- Reorder Management ----------

@app.route('/reorder-management.php')
@app.route('/reorder-management')
@login_required
def reorder_management():
    conn = get_db()
    with conn.cursor() as cur:
        cur.execute(
            "SELECT i.*,(i.reorder_level - i.quantity_on_hand) as shortage,"
            "(i.reorder_level * 2) as suggested_order_qty "
            "FROM items i WHERE i.status='Active' AND i.quantity_on_hand<=i.reorder_level "
            "ORDER BY CASE WHEN i.quantity_on_hand=0 THEN 1 "
            "WHEN i.quantity_on_hand<i.reorder_level THEN 2 ELSE 3 END, i.quantity_on_hand ASC"
        )
        low_stock_items = cur.fetchall()
    conn.close()

    out_of_stock = sum(1 for i in low_stock_items if i['quantity_on_hand'] == 0)
    critical = sum(1 for i in low_stock_items if i['quantity_on_hand'] > 0 and i['quantity_on_hand'] < (i['reorder_level'] / 2))

    return render_template('reorder_management.html',
                           low_stock_items=low_stock_items,
                           total_low_stock=len(low_stock_items),
                           out_of_stock=out_of_stock,
                           critical_stock=critical)


# ---------- Products (session-based) ----------

@app.route('/products.php', methods=['GET', 'POST'])
@app.route('/products', methods=['GET', 'POST'])
@login_required
def products():
    if 'products' not in session:
        session['products'] = []

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add_product':
            prods = session['products']
            prods.append({
                'id': len(prods) + 1,
                'name': request.form.get('product_name', ''),
                'description': request.form.get('description', ''),
                'category': request.form.get('category', ''),
                'price': request.form.get('price', 0),
                'date_issued': request.form.get('date_issued', date.today().isoformat()),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            })
            session['products'] = prods
            flash('Product added successfully!', 'success')
        elif action == 'delete_product':
            pid = int(request.form.get('id', 0))
            session['products'] = [p for p in session['products'] if p['id'] != pid]
            flash('Product deleted successfully!', 'success')
        elif action == 'upload_excel':
            f = request.files.get('excel_file')
            if f:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(f.read()))
                ws = wb.active
                imported = 0
                prods = session['products']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    prods.append({
                        'id': len(prods) + 1,
                        'name': str(row[0]),
                        'description': '',
                        'category': '',
                        'price': row[1] if len(row) > 1 else 0,
                        'date_issued': str(row[2]) if len(row) > 2 else date.today().isoformat(),
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    })
                    imported += 1
                session['products'] = prods
                flash(f'Successfully imported {imported} products from Excel file!', 'success')
        return redirect(url_for('products'))

    return render_template('products.html', products=session['products'])


# ---------- Cost (session-based) ----------

@app.route('/cost.php', methods=['GET', 'POST'])
@app.route('/cost', methods=['GET', 'POST'])
@login_required
def cost():
    if 'costs' not in session:
        session['costs'] = []

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add_cost':
            costs = session['costs']
            costs.append({
                'id': len(costs) + 1,
                'product_name': request.form.get('product_name', ''),
                'unit_cost': request.form.get('unit_cost', ''),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            })
            session['costs'] = costs
            flash('Cost added successfully!', 'success')
        elif action == 'delete_cost':
            cid = int(request.form.get('id', 0))
            session['costs'] = [c for c in session['costs'] if c['id'] != cid]
            flash('Cost deleted successfully!', 'success')
        return redirect(url_for('cost'))

    return render_template('cost.html', costs=session['costs'])


# ---------- Quantity List (session-based) ----------

@app.route('/quantity-list.php', methods=['GET', 'POST'])
@app.route('/quantity-list', methods=['GET', 'POST'])
@login_required
def quantity_list():
    if 'quantities' not in session:
        session['quantities'] = []

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add_quantity':
            qs = session['quantities']
            qs.append({
                'id': len(qs) + 1,
                'product_name': request.form.get('product_name', ''),
                'quantity': request.form.get('quantity', ''),
                'unit': request.form.get('unit', ''),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            })
            session['quantities'] = qs
            flash('Quantity added successfully!', 'success')
        elif action == 'delete_quantity':
            qid = int(request.form.get('id', 0))
            session['quantities'] = [q for q in session['quantities'] if q['id'] != qid]
            flash('Quantity deleted successfully!', 'success')
        return redirect(url_for('quantity_list'))

    return render_template('quantity_list.html', quantities=session['quantities'])


# ---------- Quantity Issued (session-based) ----------

@app.route('/quantity-issued.php', methods=['GET', 'POST'])
@app.route('/quantity-issued', methods=['GET', 'POST'])
@login_required
def quantity_issued():
    if 'issued' not in session:
        session['issued'] = []

    if request.method == 'POST':
        action = request.form.get('action', '')
        if action == 'add_issued':
            iss = session['issued']
            iss.append({
                'id': len(iss) + 1,
                'product_name': request.form.get('product_name', ''),
                'quantity': request.form.get('quantity', ''),
                'issued_to': request.form.get('issued_to', ''),
                'date_issued': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            })
            session['issued'] = iss
            flash('Quantity issued successfully!', 'success')
        elif action == 'delete_issued':
            iid = int(request.form.get('id', 0))
            session['issued'] = [i for i in session['issued'] if i['id'] != iid]
            flash('Record deleted successfully!', 'success')
        return redirect(url_for('quantity_issued'))

    return render_template('quantity_issued.html', issued=session['issued'])


# ─── Run ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app.run(debug=True, port=5000)

