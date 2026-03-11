"""Import Engineering Office items and assign to Engineering office"""
import pymysql
import openpyxl
import os
from datetime import datetime, date
from dotenv import load_dotenv

load_dotenv()

DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASS = os.getenv('DB_PASS', '')
DB_NAME = os.getenv('DB_NAME', 'mabini_inventory')

file_path = r'c:\Users\ramir\Downloads\Engineering_Office_Items.xlsx'

try:
    conn = pymysql.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASS,
        database=DB_NAME,
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )
    
    # Get Engineering office ID
    with conn.cursor() as cur:
        cur.execute("SELECT id, office_name FROM offices WHERE office_name = 'ENGINEERING'")
        engineering_office = cur.fetchone()
        
        if not engineering_office:
            print("❌ Engineering office not found!")
            conn.close()
            exit(1)
        
        engineering_id = engineering_office['id']
        print(f"✅ Found Engineering office (ID: {engineering_id})")
    
    # Load Excel file
    print(f"\n📄 Loading: {os.path.basename(file_path)}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Read headers
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    print(f"   Headers: {headers}")
    
    # Map column indices
    col_map = {header.lower(): i for i, header in enumerate(headers)}
    
    imported = 0
    updated = 0
    skipped = 0
    category_cache = {}
    
    with conn.cursor() as cur:
        # Pre-load categories
        cur.execute("SELECT id, category_name FROM categories")
        for cat in cur.fetchall():
            category_cache[cat['category_name'].strip().lower()] = cat['id']
        
        # Get max item code
        cur.execute("SELECT item_code FROM items WHERE item_code LIKE 'ENG-%' ORDER BY item_code DESC LIMIT 1")
        last_code = cur.fetchone()
        auto_num = 1
        if last_code:
            try:
                auto_num = int(last_code['item_code'].split('-')[1]) + 1
            except:
                pass
        
        print("\n" + "=" * 80)
        print("Processing items:")
        print("=" * 80)
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 3:
                skipped += 1
                continue
            
            # Extract data
            item_code_idx = col_map.get('item code', 1)
            item_name_idx = col_map.get('item name', 2)
            category_idx = col_map.get('category', 3)
            unit_idx = col_map.get('unit', 5)
            qty_idx = col_map.get('quantity', 6)
            cost_idx = col_map.get('unit cost', 7)
            date_idx = col_map.get('date acquired', 9)
            
            item_code = str(row[item_code_idx]).strip() if len(row) > item_code_idx and row[item_code_idx] else ''
            item_name = str(row[item_name_idx]).strip() if len(row) > item_name_idx and row[item_name_idx] else ''
            
            if not item_name or item_name == 'None':
                skipped += 1
                continue
            
            # Generate item code if empty
            if not item_code or item_code == 'None':
                item_code = f"ENG-{auto_num:03d}"
                auto_num += 1
            
            # Category
            category_id = None
            if category_idx < len(row) and row[category_idx]:
                cat_name = str(row[category_idx]).strip()
                key = cat_name.lower()
                if key in category_cache:
                    category_id = category_cache[key]
                else:
                    # Create new category
                    cur.execute(
                        "INSERT INTO categories (category_name, description, status) VALUES (%s, %s, 'Active')",
                        (cat_name, 'Auto-created from import')
                    )
                    conn.commit()
                    category_id = cur.lastrowid
                    category_cache[key] = category_id
            
            # Unit
            unit = 'piece'
            if unit_idx < len(row) and row[unit_idx]:
                unit = str(row[unit_idx]).strip().lower()
            
            # Quantity
            quantity = 0
            if qty_idx < len(row) and row[qty_idx]:
                try:
                    quantity = int(float(str(row[qty_idx])))
                except:
                    quantity = 0
            
            # Unit cost
            unit_cost = 0.0
            if cost_idx < len(row) and row[cost_idx]:
                try:
                    unit_cost = float(row[cost_idx])
                except:
                    unit_cost = 0.0
            
            # Date acquired
            date_acquired = None
            if date_idx < len(row) and row[date_idx]:
                if isinstance(row[date_idx], (datetime, date)):
                    date_acquired = row[date_idx].strftime('%Y-%m-%d')
                else:
                    date_acquired = str(row[date_idx]).strip() or None
            
            # Check if item exists
            cur.execute("SELECT id, item_name, quantity_on_hand FROM items WHERE item_code = %s", (item_code,))
            existing = cur.fetchone()
            
            if existing:
                # Update existing item - assign to Engineering office
                cur.execute("""
                    UPDATE items 
                    SET item_name=%s, category_id=%s, unit=%s, unit_cost=%s, 
                        quantity_on_hand=%s, date_acquired=%s, office_id=%s, updated_at=NOW()
                    WHERE item_code=%s
                """, (item_name, category_id, unit, unit_cost, quantity, date_acquired, 
                      engineering_id, item_code))
                updated += 1
                print(f"  ✏️  Updated: {item_code} - {item_name} (Qty: {quantity})")
            else:
                # Insert new item
                try:
                    cur.execute("""
                        INSERT INTO items 
                        (item_code, item_name, category_id, unit, unit_cost, quantity_on_hand,
                         reorder_level, date_acquired, office_id, status, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s, 10, %s, %s, 'Active', 1)
                    """, (item_code, item_name, category_id, unit, unit_cost, quantity, 
                          date_acquired, engineering_id))
                    imported += 1
                    print(f"  ✅ Added: {item_code} - {item_name} (Qty: {quantity})")
                except pymysql.err.IntegrityError as e:
                    print(f"  ⚠️  Duplicate: {item_code} - {item_name}")
                    skipped += 1
        
        conn.commit()
    
    conn.close()
    
    print("\n" + "=" * 80)
    print("Import Summary:")
    print("=" * 80)
    print(f"  New items added:     {imported}")
    print(f"  Existing updated:    {updated}")
    print(f"  Skipped:             {skipped}")
    print(f"  Total processed:     {imported + updated}")
    print("=" * 80)
    
    if imported > 0 or updated > 0:
        print(f"\n✅ Successfully integrated {imported + updated} Engineering Office items!")
        print(f"   All items assigned to: {engineering_office['office_name']} (Office ID: {engineering_id})")
    
except FileNotFoundError:
    print(f"❌ File not found: {file_path}")
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
