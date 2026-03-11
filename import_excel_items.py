"""Import items from Excel files"""
import pymysql
import openpyxl
import os
from datetime import datetime, date
from dotenv import load_dotenv

load_dotenv()

DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASS = os.getenv('DB_PASS', '')
DB_NAME = os.getenv('DB_NAME', 'mabini_inventory')

excel_files = [
    r'c:\Users\ramir\Downloads\inventory_template.xlsx',
    r'c:\Users\ramir\Downloads\inventory_import_template.xlsx',
    r'c:\Users\ramir\Downloads\inventory_items_20260311_103639.xlsx',
    r'c:\Users\ramir\Downloads\Engineering_Office_Items (1).xlsx',
    r'c:\Users\ramir\Downloads\Engineering_Office_Items.xlsx',
]

def import_from_excel(file_path, conn, user_id=1):
    """Import items from an Excel file"""
    if not os.path.exists(file_path):
        print(f"⚠️  File not found: {file_path}")
        return 0, 0, 0
    
    print(f"\n📄 Processing: {os.path.basename(file_path)}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Show first few rows to understand structure
        print(f"   Worksheet: {ws.title}")
        print(f"   Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Read header row
        headers = [cell.value for cell in ws[1]]
        print(f"   Headers: {headers[:5]}...")
        
        imported = 0
        skipped = 0
        duplicates = 0
        
        category_cache = {}
        office_cache = {}
        
        with conn.cursor() as cur:
            # Pre-load categories
            cur.execute("SELECT id, category_name FROM categories")
            for cat in cur.fetchall():
                category_cache[cat['category_name'].strip().lower()] = cat['id']
            
            # Pre-load offices
            cur.execute("SELECT id, office_name FROM offices WHERE status='Active'")
            for office in cur.fetchall():
                office_cache[office['office_name'].strip().lower()] = office['id']
            
            # Get max item code number
            cur.execute("SELECT item_code FROM items WHERE item_code LIKE 'ITEM-%' ORDER BY item_code DESC LIMIT 1")
            last_auto = cur.fetchone()
            auto_num = 1
            if last_auto:
                try:
                    auto_num = int(last_auto['item_code'].split('-')[1]) + 1
                except (IndexError, ValueError):
                    pass
            
            # Process rows (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 2 or (not row[0] and not row[1]):
                    skipped += 1
                    continue
                
                # Extract data (adjust indices based on actual file structure)
                item_code = str(row[0]).strip() if row[0] else ''
                item_name = str(row[1]).strip() if row[1] else ''
                
                if not item_name:
                    skipped += 1
                    continue
                
                # Auto-generate code if empty
                if not item_code:
                    item_code = f"ITEM-{auto_num:05d}"
                    auto_num += 1
                
                # Category (column 3)
                category_id = None
                if len(row) > 2 and row[2]:
                    cat_name = str(row[2]).strip()
                    key = cat_name.lower()
                    if key in category_cache:
                        category_id = category_cache[key]
                    else:
                        # Create new category
                        cur.execute(
                            "INSERT INTO categories (category_name, description, status) VALUES (%s, %s, 'Active')",
                            (cat_name, f'Auto-created from import')
                        )
                        conn.commit()
                        category_id = cur.lastrowid
                        category_cache[key] = category_id
                
                # Unit (column 4)
                unit = str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'piece'
                
                # Unit cost (column 5)
                unit_cost = 0
                if len(row) > 4 and row[4]:
                    try:
                        unit_cost = float(row[4])
                    except (ValueError, TypeError):
                        unit_cost = 0
                
                # Quantity (column 6)
                quantity = 0
                if len(row) > 5 and row[5]:
                    try:
                        quantity = int(float(str(row[5])))
                    except (ValueError, TypeError):
                        quantity = 0
                
                # Reorder level (column 7)
                reorder = 10
                if len(row) > 6 and row[6]:
                    try:
                        reorder = int(float(str(row[6])))
                    except (ValueError, TypeError):
                        reorder = 10
                
                # Expiration date (column 8)
                exp_date = None
                if len(row) > 7 and row[7]:
                    if isinstance(row[7], (datetime, date)):
                        exp_date = row[7].strftime('%Y-%m-%d')
                    else:
                        exp_str = str(row[7]).strip()
                        if exp_str:
                            exp_date = exp_str
                
                # Date acquired (column 9)
                acq_date = None
                if len(row) > 8 and row[8]:
                    if isinstance(row[8], (datetime, date)):
                        acq_date = row[8].strftime('%Y-%m-%d')
                    else:
                        acq_str = str(row[8]).strip()
                        if acq_str:
                            acq_date = acq_str
                
                # Office (column 10)
                office_id = None
                if len(row) > 9 and row[9]:
                    office_name = str(row[9]).strip()
                    key = office_name.lower()
                    if key in office_cache:
                        office_id = office_cache[key]
                
                # Insert item
                try:
                    cur.execute("""
                        INSERT INTO items 
                        (item_code, item_name, category_id, unit, unit_cost, quantity_on_hand,
                         reorder_level, expiration_date, date_acquired, office_id, status, created_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Active', %s)
                    """, (item_code, item_name, category_id, unit, unit_cost, quantity,
                          reorder, exp_date, acq_date, office_id, user_id))
                    imported += 1
                except pymysql.err.IntegrityError as e:
                    if 'Duplicate entry' in str(e):
                        duplicates += 1
                    else:
                        print(f"   ⚠️  Row {row_num}: {e}")
                        skipped += 1
        
        conn.commit()
        return imported, skipped, duplicates
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        return 0, 0, 0

def main():
    print("=" * 70)
    print("  Import Items from Excel Files")
    print("=" * 70)
    
    try:
        conn = pymysql.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASS,
            database=DB_NAME,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        
        total_imported = 0
        total_skipped = 0
        total_duplicates = 0
        files_processed = 0
        
        for file_path in excel_files:
            imported, skipped, duplicates = import_from_excel(file_path, conn)
            total_imported += imported
            total_skipped += skipped
            total_duplicates += duplicates
            if imported > 0 or skipped > 0 or duplicates > 0:
                files_processed += 1
                print(f"   ✅ Imported: {imported} | Skipped: {skipped} | Duplicates: {duplicates}")
        
        conn.close()
        
        print("\n" + "=" * 70)
        print(f"Summary:")
        print(f"  Files processed: {files_processed}")
        print(f"  Total imported: {total_imported}")
        print(f"  Total skipped: {total_skipped}")
        print(f"  Total duplicates: {total_duplicates}")
        print("=" * 70)
        
        if total_imported > 0:
            print(f"\n✅ Successfully imported {total_imported} items!")
        
    except Exception as e:
        print(f"❌ Database error: {e}")

if __name__ == '__main__':
    main()
