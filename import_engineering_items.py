"""Import items from Engineering Office Excel files with correct column mapping"""
import pymysql
import openpyxl
import os
from datetime import datetime, date
from dotenv import load_dotenv

load_dotenv()

DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASS = os.getenv('DB_PASS', '')
DB_NAME = os.getenv('DB_NAME', 'mabini_inventory')

engineering_files = [
    r'c:\Users\ramir\Downloads\Engineering_Office_Items (1).xlsx',
    r'c:\Users\ramir\Downloads\Engineering_Office_Items.xlsx',
]

def import_engineering_items(file_path, conn, user_id=1):
    """Import items from Engineering Office Excel files"""
    if not os.path.exists(file_path):
        print(f"⚠️  File not found: {file_path}")
        return 0, 0, 0
    
    print(f"\n📄 Processing: {os.path.basename(file_path)}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Read header row to find column indices
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        print(f"   Headers: {headers}")
        
        # Map column names to indices
        col_map = {}
        for i, header in enumerate(headers):
            col_map[header.lower()] = i
        
        print(f"   Column mapping: {col_map}")
        
        imported = 0
        skipped = 0
        duplicates = 0
        updated = 0
        
        category_cache = {}
        office_cache = {}
        
        with conn.cursor() as cur:
            # Pre-load categories
            cur.execute("SELECT id, category_name FROM categories")
            for cat in cur.fetchall():
                category_cache[cat['category_name'].strip().lower()] = cat['id']
            
            # Pre-load offices
            cur.execute("SELECT id, office_name FROM offices WHERE status='Active'")
            for office in cur.fetchall():
                office_cache[office['office_name'].strip().lower()] = office['id']
            
            # Get max item code number
            cur.execute("SELECT item_code FROM items WHERE item_code LIKE 'ITEM-%' ORDER BY item_code DESC LIMIT 1")
            last_auto = cur.fetchone()
            auto_num = 1
            if last_auto:
                try:
                    auto_num = int(last_auto['item_code'].split('-')[1]) + 1
                except (IndexError, ValueError):
                    pass
            
            # Process rows (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 2:
                    skipped += 1
                    continue
                
                # Extract data using column mapping
                item_code_idx = col_map.get('item code', col_map.get('item_code', 1))
                item_name_idx = col_map.get('item name', col_map.get('item_name', 2))
                category_idx = col_map.get('category', 3)
                office_idx = col_map.get('office', -1)
                
                item_code = str(row[item_code_idx]).strip() if len(row) > item_code_idx and row[item_code_idx] else ''
                item_name = str(row[item_name_idx]).strip() if len(row) > item_name_idx and row[item_name_idx] else ''
                
                if not item_name:
                    skipped += 1
                    continue
                
                # Auto-generate code if empty
                if not item_code or item_code == 'None':
                    item_code = f"ITEM-{auto_num:05d}"
                    auto_num += 1
                
                # Category
                category_id = None
                if category_idx < len(row) and row[category_idx]:
                    cat_name = str(row[category_idx]).strip()
                    key = cat_name.lower()
                    if key in category_cache:
                        category_id = category_cache[key]
                    else:
                        # Create new category
                        cur.execute(
                            "INSERT INTO categories (category_name, description, status) VALUES (%s, %s, 'Active')",
                            (cat_name, f'Auto-created from import')
                        )
                        conn.commit()
                        category_id = cur.lastrowid
                        category_cache[key] = category_id
                
                # Office
                office_id = None
                if office_idx >= 0 and office_idx < len(row) and row[office_idx]:
                    office_name = str(row[office_idx]).strip()
                    key = office_name.lower()
                    if key in office_cache:
                        office_id = office_cache[key]
                    elif 'engineering' in office_name.lower():
                        # Find or create engineering office
                        office_id = office_cache.get('engineering')
                
                # Default values
                unit = 'piece'
                unit_cost = 0
                quantity = 1
                reorder = 10
                
                # Insert or update item
                try:
                    # Check if item exists
                    cur.execute("SELECT id FROM items WHERE item_code = %s", (item_code,))
                    existing = cur.fetchone()
                    
                    if existing:
                        # Update existing item
                        cur.execute("""
                            UPDATE items 
                            SET item_name=%s, category_id=%s, office_id=%s, updated_at=NOW()
                            WHERE item_code=%s
                        """, (item_name, category_id, office_id, item_code))
                        updated += 1
                    else:
                        # Insert new item
                        cur.execute("""
                            INSERT INTO items 
                            (item_code, item_name, category_id, unit, unit_cost, quantity_on_hand,
                             reorder_level, office_id, status, created_by)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Active', %s)
                        """, (item_code, item_name, category_id, unit, unit_cost, quantity,
                              reorder, office_id, user_id))
                        imported += 1
                        
                except pymysql.err.IntegrityError as e:
                    if 'Duplicate entry' in str(e):
                        duplicates += 1
                    else:
                        print(f"   ⚠️  Row {row_num}: {e}")
                        skipped += 1
        
        conn.commit()
        return imported, updated, skipped, duplicates
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 0, 0, 0, 0

def main():
    print("=" * 70)
    print("  Import Engineering Office Items")
    print("=" * 70)
    
    try:
        conn = pymysql.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASS,
            database=DB_NAME,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        
        total_imported = 0
        total_updated = 0
        total_skipped = 0
        total_duplicates = 0
        files_processed = 0
        
        for file_path in engineering_files:
            imported, updated, skipped, duplicates = import_engineering_items(file_path, conn)
            total_imported += imported
            total_updated += updated
            total_skipped += skipped
            total_duplicates += duplicates
            if imported > 0 or updated > 0 or skipped > 0 or duplicates > 0:
                files_processed += 1
                print(f"   ✅ New: {imported} | Updated: {updated} | Skipped: {skipped} | Duplicates: {duplicates}")
        
        conn.close()
        
        print("\n" + "=" * 70)
        print(f"Summary:")
        print(f"  Files processed: {files_processed}")
        print(f"  Total new items: {total_imported}")
        print(f"  Total updated: {total_updated}")
        print(f"  Total skipped: {total_skipped}")
        print(f"  Total duplicates: {total_duplicates}")
        print("=" * 70)
        
        if total_imported > 0 or total_updated > 0:
            print(f"\n✅ Import complete!")
        
    except Exception as e:
        print(f"❌ Database error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
